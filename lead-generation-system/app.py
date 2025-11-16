from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from datetime import datetime
import sqlite3
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)  # Enable CORS for frontend

# Database configuration
DATABASE = 'leads.db'

# ===== Database Functions =====
def get_db_connection():
    """Create and return a database connection"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    """Initialize the database with the leads table"""
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            first_name TEXT NOT NULL,
            last_name TEXT NOT NULL,
            email TEXT NOT NULL,
            phone TEXT NOT NULL,
            company TEXT,
            job_title TEXT,
            industry TEXT,
            lead_source TEXT NOT NULL,
            address TEXT,
            city TEXT,
            state TEXT,
            country TEXT,
            zip_code TEXT,
            notes TEXT,
            status TEXT NOT NULL DEFAULT 'New',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()
    print("Database initialized successfully")

# ===== API Routes =====

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({'status': 'healthy', 'message': 'API is running'}), 200

@app.route('/api/leads', methods=['GET'])
def get_leads():
    """Get all leads"""
    try:
        conn = get_db_connection()
        leads = conn.execute('''
            SELECT * FROM leads ORDER BY created_at DESC
        ''').fetchall()
        conn.close()
        
        leads_list = [dict(lead) for lead in leads]
        return jsonify(leads_list), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/leads', methods=['POST'])
def create_lead():
    """Create a new lead"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['first_name', 'last_name', 'email', 'phone', 'lead_source', 'status']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'{field} is required'}), 400
        
        conn = get_db_connection()
        conn.execute('''
            INSERT INTO leads (
                first_name, last_name, email, phone, company, job_title,
                industry, lead_source, address, city, state, country,
                zip_code, notes, status, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['first_name'],
            data['last_name'],
            data['email'],
            data['phone'],
            data.get('company', ''),
            data.get('job_title', ''),
            data.get('industry', ''),
            data['lead_source'],
            data.get('address', ''),
            data.get('city', ''),
            data.get('state', ''),
            data.get('country', ''),
            data.get('zipCode', ''),
            data.get('notes', ''),
            data['status'],
            datetime.now().isoformat(),
            datetime.now().isoformat()
        ))
        conn.commit()
        lead_id = conn.lastrowid
        conn.close()
        
        return jsonify({'id': lead_id, 'message': 'Lead created successfully'}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/leads/<int:lead_id>', methods=['GET'])
def get_lead(lead_id):
    """Get a specific lead by ID"""
    try:
        conn = get_db_connection()
        lead = conn.execute('SELECT * FROM leads WHERE id = ?', (lead_id,)).fetchone()
        conn.close()
        
        if lead:
            return jsonify(dict(lead)), 200
        else:
            return jsonify({'error': 'Lead not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/leads/<int:lead_id>', methods=['PUT'])
def update_lead(lead_id):
    """Update a lead"""
    try:
        data = request.get_json()
        
        conn = get_db_connection()
        
        # Check if lead exists
        lead = conn.execute('SELECT * FROM leads WHERE id = ?', (lead_id,)).fetchone()
        if not lead:
            conn.close()
            return jsonify({'error': 'Lead not found'}), 404
        
        # Update lead
        conn.execute('''
            UPDATE leads SET
                first_name = ?, last_name = ?, email = ?, phone = ?,
                company = ?, job_title = ?, industry = ?, lead_source = ?,
                address = ?, city = ?, state = ?, country = ?,
                zip_code = ?, notes = ?, status = ?, updated_at = ?
            WHERE id = ?
        ''', (
            data.get('firstName', lead['first_name']),
            data.get('lastName', lead['last_name']),
            data.get('email', lead['email']),
            data.get('phone', lead['phone']),
            data.get('company', lead['company']),
            data.get('jobTitle', lead['job_title']),
            data.get('industry', lead['industry']),
            data.get('leadSource', lead['lead_source']),
            data.get('address', lead['address']),
            data.get('city', lead['city']),
            data.get('state', lead['state']),
            data.get('country', lead['country']),
            data.get('zipCode', lead['zip_code']),
            data.get('notes', lead['notes']),
            data.get('status', lead['status']),
            datetime.now().isoformat(),
            lead_id
        ))
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Lead updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/leads/<int:lead_id>', methods=['DELETE'])
def delete_lead(lead_id):
    """Delete a lead"""
    try:
        conn = get_db_connection()
        
        # Check if lead exists
        lead = conn.execute('SELECT * FROM leads WHERE id = ?', (lead_id,)).fetchone()
        if not lead:
            conn.close()
            return jsonify({'error': 'Lead not found'}), 404
        
        conn.execute('DELETE FROM leads WHERE id = ?', (lead_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'message': 'Lead deleted successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/leads/export', methods=['GET'])
def export_leads():
    """Export all leads to Excel"""
    try:
        conn = get_db_connection()
        leads = conn.execute('SELECT * FROM leads ORDER BY created_at DESC').fetchall()
        conn.close()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"
        
        # Define headers
        headers = [
            'ID', 'First Name', 'Last Name', 'Email', 'Phone', 'Company',
            'Job Title', 'Industry', 'Lead Source', 'Address', 'City',
            'State', 'Country', 'Zip Code', 'Notes', 'Status', 'Created At', 'Updated At'
        ]
        
        # Style header row
        header_fill = PatternFill(start_color="6366f1", end_color="6366f1", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_num, lead in enumerate(leads, 2):
            ws.cell(row=row_num, column=1, value=lead['id'])
            ws.cell(row=row_num, column=2, value=lead['first_name'])
            ws.cell(row=row_num, column=3, value=lead['last_name'])
            ws.cell(row=row_num, column=4, value=lead['email'])
            ws.cell(row=row_num, column=5, value=lead['phone'])
            ws.cell(row=row_num, column=6, value=lead['company'] or '')
            ws.cell(row=row_num, column=7, value=lead['job_title'] or '')
            ws.cell(row=row_num, column=8, value=lead['industry'] or '')
            ws.cell(row=row_num, column=9, value=lead['lead_source'])
            ws.cell(row=row_num, column=10, value=lead['address'] or '')
            ws.cell(row=row_num, column=11, value=lead['city'] or '')
            ws.cell(row=row_num, column=12, value=lead['state'] or '')
            ws.cell(row=row_num, column=13, value=lead['country'] or '')
            ws.cell(row=row_num, column=14, value=lead['zip_code'] or '')
            ws.cell(row=row_num, column=15, value=lead['notes'] or '')
            ws.cell(row=row_num, column=16, value=lead['status'])
            ws.cell(row=row_num, column=17, value=lead['created_at'])
            ws.cell(row=row_num, column=18, value=lead['updated_at'])
        
        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            max_length = max(
                len(str(header)),
                max((len(str(ws.cell(row=row_num, column=col_num).value)) 
                     for row_num in range(2, len(leads) + 2) if ws.cell(row=row_num, column=col_num).value), 
                    default=0)
            )
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to temporary file
        filename = f'leads_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(filename)
        
        return send_file(filename, as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ===== Main =====
if __name__ == '__main__':
    # Initialize database
    init_db()
    
    # Run the Flask app
    print("Starting Lead Generation System API...")
    print("API will be available at http://localhost:5000")
    print("Frontend should be opened in a browser")
    app.run(debug=True, host='0.0.0.0', port=5000)

